from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Credito, PagoCredito, AprobacionCredito, CuotaCredito, ServicioCredito,SolicitudCredito, NombrePrendas, ReciboPago
from .forms import CreditoForm, PagoForm, DocumentosUsuarioForm, RenovarServicioForm, UsuarioCreditoForm, VehiculoCreditoForm, NombrePrendasForm
from datetime import date, datetime, timedelta
from django.http import JsonResponse
from django import forms
from django.forms import formset_factory
from decimal import Decimal
from weasyprint import HTML
from django.template.loader import render_to_string
from django.http import HttpResponse, FileResponse
from django.conf import settings
import os
from contratos.models import usuario, Vehiculo_contratos
from django.db import transaction
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from tempfile import NamedTemporaryFile
from dateutil.relativedelta import relativedelta
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth
from django.utils import timezone
from collections import defaultdict
from django.core.paginator import Paginator
from num2words import num2words
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from .utils import requiere_sede_financiera, requiere_sede_financiera_admin
from django.core.files.base import ContentFile


# ----------------------------------------------------------------------
# DETALLE DEL CRÉDITO
# ----------------------------------------------------------------------
@login_required
@requiere_sede_financiera_admin
def detalle_credito(request, credito_id):
    credito = get_object_or_404(Credito, id=credito_id)
    cronograma = credito.cuotas_credito.all().order_by('numero')
    pagos = credito.pagos.all().order_by('fecha_pago')
    prendas = NombrePrendas.objects.all()
    recibos = credito.recibos.prefetch_related("pagos")


    saldo_capital = credito.saldo_capital()
    intereses_pendientes = credito.intereses_pendientes()
    saldo_total = credito.saldo_total()

    # Servicios activos
    servicios_activos = credito.servicios.filter(estado='Activo')

    # -------------------------------
    # CÁLCULO DE MORATORIOS
    # -------------------------------
    hoy = timezone.now().date()

    cuotas_mora_info = []
    total_interes_moratorio = Decimal('0.00')
    total_cobro_juridico = Decimal('0.00')
    total_general_moratorios = Decimal('0.00')

    for cuota in cronograma:
        info = credito.calcular_moratorios_por_cuota(cuota, hoy=hoy)

        total_interes_moratorio += info['interes_moratorio']
        total_cobro_juridico += info['cobro_juridico']
        total_general_moratorios += info['total_moratorios']

        cuotas_mora_info.append({
            'cuota': cuota,
            'moratorio': info
        })

    moratorios_acumulados = {
        'total_interes_moratorio': total_interes_moratorio.quantize(Decimal('0.01')),
        'total_cobro_juridico': total_cobro_juridico.quantize(Decimal('0.01')),
        'total_general': total_general_moratorios.quantize(Decimal('0.01')),
    }

    # -------------------------------
    # RENDER
    # -------------------------------
    return render(request, 'creditos/detalle_credito.html', {
        'credito': credito,
        'cronograma': cronograma,
        'pagos': pagos,
        'saldo_capital': saldo_capital,
        'intereses_pendientes': intereses_pendientes,
        'saldo': saldo_total,
        'servicios_activos': servicios_activos,

        # Nuevos datos para el template
        'cuotas_mora_info': cuotas_mora_info,
        'moratorios_acumulados': moratorios_acumulados,
        'prendas': prendas,
        'recibos': recibos,
    })

# ----------------------------------------------------------------------
# AJAX: CALCULAR INTERÉS MENSUAL
# ----------------------------------------------------------------------
def calcular_interes_ajax(request, credito_id):
    credito = get_object_or_404(Credito, pk=credito_id)
    interes_mes = credito.saldo_capital() * (credito.interes / 100)
    return JsonResponse({'interes_mes': float(interes_mes)})


# ----------------------------------------------------------------------
# CREAR NUEVO CRÉDITO
# ----------------------------------------------------------------------
@login_required
@requiere_sede_financiera_admin
def crear_credito(request):
    if request.method == 'POST':
        form = CreditoForm(request.POST)

        if form.is_valid():

            # ===============================
            # 1. Leer cédula y placa digitadas
            # ===============================
            cedula = form.cleaned_data["cedula"]
            placa = form.cleaned_data["placa"]

            # ===============================
            # 2. Buscar usuario
            # ===============================
            try:
                usuario_obj = usuario.objects.get(cedula=cedula)
            except usuario.DoesNotExist:
                messages.error(request, f"No existe un cliente con cédula {cedula}.")
                return render(request, "creditos/crear_credito.html", {"form": form})

            # ===============================
            # 3. Buscar vehículo
            # ===============================
            try:
                vehiculo_obj = Vehiculo_contratos.objects.get(placa__iexact=placa)
            except Vehiculo_contratos.DoesNotExist:
                messages.error(request, f"No existe un vehículo con placa {placa}.")
                return render(request, "creditos/crear_credito.html", {"form": form})

            # ===============================
            # 4. Crear el crédito
            # ===============================
            credito = form.save(commit=False)
            credito.usuario = usuario_obj
            credito.vehiculo = vehiculo_obj
            credito.save()

            # ===============================
            # 5. Procesar Servicios (Seguro / GPS)
            # ===============================
            seguro_valor = form.cleaned_data.get("seguro_valor")
            gps_valor = form.cleaned_data.get("gps_valor")
            inicio = credito.fecha_inicio

            # Fecha fin obligatoria (12 meses menos 1 día)
            fin = inicio + relativedelta(months=12) - relativedelta(days=1)

            # ---- Seguro Primer Año ----
            if seguro_valor:
                ServicioCredito.objects.update_or_create(
                    credito=credito,
                    tipo=ServicioCredito.TIPO_SEG,
                    fecha_inicio=inicio,
                    defaults={
                        "valor_anual": seguro_valor,
                        "estado": ServicioCredito.ESTADO_ACTIVO,
                        "fecha_fin": fin,           # 🔥 IMPORTANTE
                    }
                )

            # ---- GPS Primer Año ----
            if gps_valor:
                ServicioCredito.objects.update_or_create(
                    credito=credito,
                    tipo=ServicioCredito.TIPO_GPS,
                    fecha_inicio=inicio,
                    defaults={
                        "valor_anual": gps_valor,
                        "estado": ServicioCredito.ESTADO_ACTIVO,
                        "fecha_fin": fin,           # 🔥 IMPORTANTE
                    }
                )

            # ===============================
            # 6. Generar cronograma
            # ===============================
            try:
                credito.generar_cronograma()
            except ValueError as e:
                messages.warning(request, f"Crédito creado pero faltan renovaciones: {e}")
                return redirect('detalle_credito', credito_id=credito.id)

            # ===============================
            # 7. Éxito total
            # ===============================
            messages.success(request, f"Crédito creado correctamente para {usuario_obj.nombre}.")
            return redirect('detalle_credito', credito_id=credito.id)

    else:
        form = CreditoForm()

    return render(request, 'creditos/crear_credito.html', {'form': form})

# ----------------------------------------------------------------------
# REGISTRAR PAGO O ABONO A CAPITAL
# ----------------------------------------------------------------------
@login_required
@requiere_sede_financiera_admin
def registrar_pago(request, credito_id):
    credito = get_object_or_404(Credito, id=credito_id)

    # Mostrar solo cuotas no pagadas
    cuotas_disponibles = credito.cuotas_credito.filter(pagada=False).order_by('numero')
    opciones_cuotas = [
        (c.numero, f"Cuota {c.numero} - vence {c.fecha_vencimiento} - ${c.cuota_total:,.2f}")
        for c in cuotas_disponibles
    ]

    PagoFormSet = formset_factory(PagoForm, extra=1, can_delete=True)

    # ==========================================================
    # POST
    # ==========================================================
    if request.method == 'POST':
        formset = PagoFormSet(request.POST)

        if formset.is_valid():
            tipo_operacion = request.POST.get('accion')
            pagos_creados_ids = []
            hoy = date.today()

            # ==========================================================
            # 🟥  PAGO COMPLETO AL DÍA (INTERÉS REAL)
            # ==========================================================
            if tipo_operacion == "pago_total":

                saldo_capital = credito.saldo_capital()
                interes_real_info = credito.intereses_al_dia()
                interes_real = interes_real_info["interes_generado"]
                moratorios = credito.moratorios_totales()

                total_pagar_hoy = (
                    saldo_capital +
                    interes_real +
                    moratorios['total_general']
                ).quantize(Decimal("0.01"))

                # Registrar pago completo final
                pago = PagoCredito.objects.create(
                    credito=credito,
                    valor_pagado=saldo_capital,
                    interes_pagado=interes_real,
                    interes_moratorio_pagado=moratorios["total_interes_moratorio"],
                    cobro_juridico_pagado=moratorios["total_cobro_juridico"],
                    tipo_pago='cuota',
                    observacion='Pago total del crédito al día'
                )

                pagos_creados_ids.append(pago.id)

                # Marcar todas las cuotas como pagadas
                credito.cuotas_credito.update(pagada=True, estado="Pagada")

                # Finalizar crédito
                credito.estado = "Finalizado"
                credito.save()

                # AJAX
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    url = reverse('generar_recibo') + f"?ids={pago.id}"
                    return JsonResponse({
                        "ok": True,
                        "recibo_url": url,
                        "redirect_url": reverse('detalle_credito', args=[credito.id])
                    })

                messages.success(request, "Pago total del crédito registrado correctamente.")
                return redirect('detalle_credito', credito_id=credito.id)

            # ==========================================================
            # 🟩 ABONO DIRECTO A CAPITAL
            # ==========================================================
            if tipo_operacion == 'abono':
                total_abono = sum(
                    Decimal(form.cleaned_data.get('valor_pagado') or 0)
                    for form in formset if form.cleaned_data and not form.cleaned_data.get('DELETE')
                )

                if total_abono > 0:
                    credito.aplicar_abono_capital(total_abono)
                    pago = PagoCredito.objects.create(
                        credito=credito,
                        valor_pagado=total_abono,
                        tipo_pago='abono',
                        observacion='Abono directo a capital'
                    )
                    pagos_creados_ids.append(pago.id)

                credito.verificar_estado_credito()

                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    url = reverse('generar_recibo') + f"?ids={','.join(map(str, pagos_creados_ids))}"
                    return JsonResponse({
                        "ok": True,
                        "recibo_url": url,
                        "redirect_url": reverse('detalle_credito', args=[credito.id])
                    })

                messages.success(request, f"Abono de ${total_abono:,.2f} aplicado correctamente.")
                return redirect('detalle_credito', credito_id=credito.id)

            # ==========================================================
            # 🟦 PAGO NORMAL DE CUOTAS
            # ==========================================================
            for form in formset:
                if not form.cleaned_data or form.cleaned_data.get('DELETE'):
                    continue

                cuota_numero = form.cleaned_data.get('cuota_numero')
                valor_pagado = Decimal(form.cleaned_data.get('valor_pagado') or 0)
                observacion = form.cleaned_data.get('observacion', '')

                interes_mora_pagado = Decimal(
                    form.cleaned_data.get('interes_moratorio_pagado')
                    or request.POST.get(f"form-{form.prefix.split('-')[1]}-mora")
                    or 0
                )

                cobro_juridico_pagado = Decimal(
                    form.cleaned_data.get('cobro_juridico_pagado')
                    or request.POST.get(f"form-{form.prefix.split('-')[1]}-juridico")
                    or 0
                )

                if cuota_numero:
                    cuota = credito.cuotas_credito.filter(
                        numero=cuota_numero,
                        pagada=False
                    ).first()

                    if cuota:

                        # Cálculo real mora
                        mora_info = credito.calcular_moratorios_por_cuota(cuota, hoy=hoy)

                        if interes_mora_pagado == 0:
                            interes_mora_pagado = mora_info['interes_moratorio']

                        if cobro_juridico_pagado == 0:
                            cobro_juridico_pagado = mora_info['cobro_juridico']

                        valor_total_cuota = (
                            cuota.cuota_total +
                            interes_mora_pagado +
                            cobro_juridico_pagado
                        ).quantize(Decimal("0.01"))

                        pago_real = (
                            valor_pagado +
                            interes_mora_pagado +
                            cobro_juridico_pagado
                        ).quantize(Decimal("0.01"))

                        # Registrar pago principal
                        pago = PagoCredito.objects.create(
                            credito=credito,
                            valor_pagado=valor_pagado,
                            cuota_numero=cuota.numero,
                            tipo_pago='cuota',
                            observacion=observacion,
                            interes_moratorio_pagado=interes_mora_pagado,
                            cobro_juridico_pagado=cobro_juridico_pagado
                        )
                        pagos_creados_ids.append(pago.id)

                        # Pago parcial
                        if pago_real < valor_total_cuota - Decimal("0.50"):
                            restante = valor_total_cuota - pago_real
                            cuota.cuota_total = restante
                            cuota.pagada = False
                            cuota.estado = "Pendiente"
                            cuota.save()

                        # Pago exacto
                        elif abs(pago_real - valor_total_cuota) <= Decimal("0.50"):
                            cuota.pagada = True
                            cuota.estado = "Pagada"
                            cuota.save(update_fields=['pagada', 'estado'])

                        # Pago mayor
                        else:
                            excedente = pago_real - valor_total_cuota
                            cuota.pagada = True
                            cuota.estado = "Pagada"
                            cuota.save(update_fields=['pagada', 'estado'])

                            pago_exced = PagoCredito.objects.create(
                                credito=credito,
                                valor_pagado=excedente,
                                tipo_pago='abono',
                                observacion=f"Excedente pago cuota {cuota.numero}"
                            )
                            pagos_creados_ids.append(pago_exced.id)

                            credito.aplicar_abono_capital(excedente)

            # Si no hay cuotas pendientes → finalizar crédito
            if credito.cuotas_credito.filter(pagada=False).count() == 0:
                credito.estado = 'Finalizado'
                credito.save()

            credito.verificar_estado_credito()

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                url = reverse('generar_recibo') + f"?ids={','.join(map(str, pagos_creados_ids))}"
                return JsonResponse({
                    "ok": True,
                    "recibo_url": url,
                    "redirect_url": reverse('detalle_credito', args=[credito.id])
                })

            messages.success(request, "Pagos registrados correctamente.")
            return redirect('detalle_credito', credito_id=credito.id)

    # ==========================================================
    # GET
    # ==========================================================
    else:
        formset = PagoFormSet()
        for f in formset:
            f.fields['cuota_numero'].widget = forms.Select(choices=opciones_cuotas)

    # -------------------------------
    # VALORES PARA "PAGO COMPLETO AL DÍA"
    # -------------------------------
    saldo_capital = credito.saldo_capital()
    interes_al_dia = credito.intereses_al_dia()
    interes_real = interes_al_dia["interes_generado"]
    moratorios = credito.moratorios_totales()

    total_pagar_hoy = (
        saldo_capital +
        interes_real +
        moratorios["total_general"]
    ).quantize(Decimal("0.01"))

    return render(request, 'creditos/registrar_pago_multiple.html', {
        'credito': credito,
        'formset': formset,
        'saldo': saldo_capital,
        'interes_real': interes_real,
        'moratorios': moratorios,
        'total_pagar_hoy': total_pagar_hoy
    })

# ----------------------------------------------------------------------
# AJAX: OBTENER VALOR DE UNA CUOTA
# ----------------------------------------------------------------------
@login_required
def obtener_valor_cuota(request, credito_id):
    credito = get_object_or_404(Credito, id=credito_id)
    cuota_num = request.GET.get("cuota")

    cuota = credito.cuotas_credito.filter(numero=cuota_num).first()
    if not cuota:
        return JsonResponse({"error": "Cuota no encontrada"}, status=404)

    # usar el cálculo EXACTO del modelo (el mismo que usa la tabla)
    mora_info = credito.calcular_moratorios_por_cuota(cuota)

    # devolver claves que el JS espera: valor_cuota, mora, juridico
    return JsonResponse({
        "valor_cuota": float(cuota.cuota_total),
        "mora": float(mora_info["interes_moratorio"]),
        "juridico": float(mora_info["cobro_juridico"]),
        "total": float(cuota.cuota_total + mora_info["total_moratorios"]),
    })


@login_required
@requiere_sede_financiera_admin
def generar_aprobacion_pdf(request):
    if request.method == 'GET' and 'cedula' in request.GET and 'placa' in request.GET:
        cedula = request.GET.get('cedula')
        placa = request.GET.get('placa')

        # Buscar el usuario y vehículo
        try:
            user = usuario.objects.get(cedula=cedula)
        except usuario.DoesNotExist:
            return HttpResponse("Error: Usuario no encontrado.", status=404)

        try:
            vehiculo = Vehiculo_contratos.objects.get(placa=placa)
        except Vehiculo_contratos.DoesNotExist:
            return HttpResponse("Error: Vehículo no encontrado.", status=404)

        # Buscar el crédito asociado
        try:
            credito = Credito.objects.filter(usuario=user, vehiculo=vehiculo).latest('id')
        except Credito.DoesNotExist:
            return HttpResponse("Error: No se encontró crédito asociado a este usuario y vehículo.", status=404)

        # Fecha y formato
        fecha_aprobacion = datetime.now().strftime("%d-%m-%Y")

        # Contexto para la plantilla
        context = {
            'credito': credito,
            'usuario': user,
            'vehiculo': vehiculo,
            'fecha_aprobacion': fecha_aprobacion,
        }

        # Renderizamos el HTML
        html_string = render_to_string('creditos/aprobacion.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri())

        # Guardar el PDF temporalmente
        pdf = html.write_pdf()
        filename = f"aprobacion_credito_{user.cedula}_{vehiculo.placa}.pdf"

        # Guardar archivo en MEDIA_ROOT (opcional)
        pdf_path = os.path.join(settings.MEDIA_ROOT, 'creditos', filename)
        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
        with open(pdf_path, 'wb') as f:
            f.write(pdf)

        # Respuesta
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    # Si no hay GET, mostrar formulario
    return render(request, 'creditos/form_aprobacion.html')

@login_required
@requiere_sede_financiera_admin
def contrato_prenda_view(request):
    if request.method == 'POST':
        cedula = request.POST.get('cedula')
        placa = request.POST.get('placa')

        try:
            user = get_object_or_404(usuario, cedula=cedula)
            vehiculo = get_object_or_404(Vehiculo_contratos, placa=placa)
        except:
            return render(request, 'creditos/contrato_prenda_form.html', {
                'error': 'No se encontró el usuario o el vehículo especificado.'
            })

        # Renderizar el HTML con los datos
        html_string = render_to_string('creditos/contrato_prenda.html', {
            'user': user,
            'vehiculo': vehiculo,
            'fecha': datetime.now()
        })

        # Generar PDF
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf_file = html.write_pdf()

        # Nombre del archivo
        filename = f"Contrato_Prenda_{vehiculo.placa}_{user.cedula}.pdf"

        # Responder el PDF directamente
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    # Si no es POST, mostrar el formulario
    return render(request, 'creditos/contrato_prenda_form.html')


# ----------------------------------------------------------------------
# SIMULADOR DE CRÉDITO (sin guardar en BD)
# ----------------------------------------------------------------------
@login_required
@requiere_sede_financiera_admin
def simulador_credito(request):

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        telefono = request.POST.get('telefono')
        valor = Decimal(request.POST.get('valor') or 0)
        cuotas = int(request.POST.get('cuotas') or 0)
        interes = Decimal(request.POST.get('interes') or 0)

        interes_mensual = interes / Decimal('100')
        saldo = valor
        cronograma = []

        for i in range(1, cuotas + 1):
            interes_mes = saldo * interes_mensual
            abono_capital = valor / Decimal(cuotas)
            cuota_total = abono_capital + interes_mes
            saldo -= abono_capital

            cronograma.append({
                'numero': i,
                'abono_capital': round(abono_capital, 2),
                'interes': round(interes_mes, 2),
                'cuota_total': round(cuota_total, 2),
                'saldo_restante': round(saldo, 2)
            })

        total_intereses = sum(c['interes'] for c in cronograma)
        total_pagar = valor + total_intereses

        context = {
            'nombre': nombre,
            'telefono': telefono,
            'valor': valor,
            'cuotas': cuotas,
            'interes': interes,
            'cronograma': cronograma,
            'total_intereses': round(total_intereses, 2),
            'total_pagar': round(total_pagar, 2),
        }

        if 'generar_pdf' in request.POST:
            html_string = render_to_string('creditos/simulacion_pdf.html', context)
            html = HTML(string=html_string, base_url=request.build_absolute_uri())
            pdf = html.write_pdf()
            filename = f"Simulacion_Credito_{nombre.replace(' ', '_')}.pdf"
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{filename}"'
            return response

        return render(request, 'creditos/simulador_credito.html', context)

    return render(request, 'creditos/simulador_credito.html')

@login_required
def generar_aprobacion_auto(request):

    if request.method == 'POST':
        cedula = request.POST.get('cedula')
        placa = request.POST.get('placa')

        try:
            user = usuario.objects.get(cedula=cedula)
            vehiculo = Vehiculo_contratos.objects.get(placa=placa)
        except Exception as e:
            return HttpResponse(f"Error al buscar datos: {e}", status=400)

        # Generar PDF
        context = {
            'usuario': user,
            'vehiculo': vehiculo,
            'fecha_aprobacion': datetime.now().strftime("%d/%m/%Y"),
        }
        html_string = render_to_string('creditos/aprobacion.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf = html.write_pdf()

        # Guardar archivo
        filename = f"Aprobacion_{user.cedula}_{vehiculo.placa}.pdf"
        folder = os.path.join(settings.MEDIA_ROOT, 'aprobaciones')
        os.makedirs(folder, exist_ok=True)
        pdf_path = os.path.join(folder, filename)
        with open(pdf_path, 'wb') as f:
            f.write(pdf)

        # Crear registro
        aprobacion = AprobacionCredito.objects.create(
            usuario=user,
            vehiculo=vehiculo,
            estado='En Estudio',
            archivo_pdf=f"aprobaciones/{filename}"
        )

        # 🔥 Devolver JSON para que JS abra el PDF y redirija
        pdf_url = settings.MEDIA_URL + f"aprobaciones/{filename}"
        return JsonResponse({'pdf_url': pdf_url, 'redirect_url': '/creditos/aprobaciones/'})

    return render(request, 'creditos/form_aprobacion.html')

@login_required
@requiere_sede_financiera_admin
def lista_aprobaciones(request):
    """
    Lista de todas las aprobaciones guardadas en la base de datos.
    """
    aprobaciones = AprobacionCredito.objects.all().order_by('-fecha')
    return render(request, 'creditos/lista_aprobaciones.html', {'aprobaciones': aprobaciones})

def generar_prenda_auto(request, aprobacion_id):
    """
    Genera la prenda del vehículo solo si la aprobación está en estado 'Aprobado'.
    """
    aprobacion = get_object_or_404(AprobacionCredito, id=aprobacion_id)

    if aprobacion.estado != 'Aprobado':
        messages.error(request, "No puedes generar la prenda hasta que la aprobación sea aprobada.")
        return redirect('lista_aprobaciones')

    user = aprobacion.usuario
    vehiculo = aprobacion.vehiculo

    context = {
        'user': user,
        'vehiculo': vehiculo,
        'fecha': datetime.now()
    }

    html_string = render_to_string('creditos/contrato_prenda.html', context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()

    filename = f"Prenda_{user.cedula}_{vehiculo.placa}.pdf"
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response

@login_required
@requiere_sede_financiera_admin
def evaluar_aprobacion(request, aprobacion_id):
    aprobacion = get_object_or_404(AprobacionCredito, id=aprobacion_id)

    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado not in ['Aprobado', 'Rechazado']:
            messages.error(request, "Estado no válido.")
            return redirect('lista_aprobaciones')

        aprobacion.estado = nuevo_estado
        aprobacion.save(update_fields=['estado'])
        messages.success(request, f"La aprobación pasó a estado: {nuevo_estado}")
        return redirect('lista_aprobaciones')

    return render(request, 'creditos/evaluar_aprobacion.html', {'aprobacion': aprobacion})

@login_required
@requiere_sede_financiera_admin
def generar_solicitud_credito(request):
    if request.method == 'POST':

        # ----- Datos del formulario -----
        cedula = request.POST.get('cedula')
        placa = request.POST.get('placa')
        valor = request.POST.get('valor')
        cuotas = request.POST.get('cuotas')
        interes = request.POST.get('interes')

        tiene_garantia = request.POST.get('tiene_garantia')
        placa_garantia = request.POST.get('placa_garantia')

        # ----- Validación de campos obligatorios -----
        if not (cedula and placa and valor and cuotas and interes):
            messages.error(request, "Todos los campos son obligatorios.")
            return redirect("generar_solicitud_credito")

        # ----- Buscar usuario principal -----
        try:
            user = usuario.objects.get(cedula=cedula)
        except usuario.DoesNotExist:
            messages.error(request, "El usuario con esa cédula no existe.")
            return redirect("generar_solicitud_credito")

        # ----- Buscar vehículo principal -----
        try:
            vehiculo = Vehiculo_contratos.objects.get(placa=placa)
        except Vehiculo_contratos.DoesNotExist:
            messages.error(request, "El vehículo con esa placa no existe.")
            return redirect("generar_solicitud_credito")

        # ----- Buscar vehículo de garantía (opcional) -----
        vehiculo_garantia = None
        if tiene_garantia and placa_garantia:
            try:
                vehiculo_garantia = Vehiculo_contratos.objects.get(placa=placa_garantia)
            except Vehiculo_contratos.DoesNotExist:
                messages.error(request, "El vehículo de garantía no existe.")
                return redirect("generar_solicitud_credito")

        # ============================================================
        #  🔥 GUARDAR LA SOLICITUD EN LA BASE DE DATOS
        # ============================================================
        solicitud = SolicitudCredito.objects.create(
            usuario=user,
            vehiculo=vehiculo,
            valor=valor,
            cuotas=cuotas,
            interes=interes,
        )

        # ============================================================
        #  CONTEXTO PARA PDF → usando SOLO LOS DATOS YA GUARDADOS
        # ============================================================
        context = {
            'p_nombre': user.nombre,
            's_nombre': getattr(user, 's_nombre', ''),
            'p_apellido': user.p_apellido,
            's_apellido': user.s_apellido,
            'cedula': user.cedula,
            'telefono': user.telefono,
            'direccion': user.direccion,

            'marca': vehiculo.marca,
            'linea': vehiculo.linea,
            'modelo': vehiculo.modelo,
            'placa': vehiculo.placa,

            'valor': valor,
            'interes': interes,
            'cuotas': cuotas,
            'id_solicitud': solicitud.id,

            # Garantía
            'garantia_marca': vehiculo_garantia.marca if vehiculo_garantia else '',
            'garantia_linea': vehiculo_garantia.linea if vehiculo_garantia else '',
            'garantia_modelo': vehiculo_garantia.modelo if vehiculo_garantia else '',
            'garantia_placa': vehiculo_garantia.placa if vehiculo_garantia else '',
        }

        # ============================================================
        #  GENERAR PDF DE LA SOLICITUD
        # ============================================================
        html_string = render_to_string('creditos/solicitud_credito_deudor.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf = html.write_pdf()

        filename = f"Solicitud_Credito_{user.cedula}.pdf"
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'

        return response

    # GET → mostrar formulario
    return render(request, 'creditos/form_solicitud_credito.html')



@login_required
def generar_excel_cronograma(request, credito_id):

    credito = get_object_or_404(Credito, id=credito_id)
    cuotas = credito.cuotas_credito.filter(pagada=False).order_by('numero')

    if not cuotas.exists():
        return HttpResponse("No hay cuotas pendientes para generar el Excel")

    # ---------------------------
    # 🔥 OBTENER PRENDA DESDE EL SELECT
    # ---------------------------
    prenda_id = request.GET.get("prenda_id")
    if not prenda_id:
        return HttpResponse("Debe seleccionar una persona de la prenda")

    prenda = get_object_or_404(NombrePrendas, id=prenda_id)

    cedula_cliente = credito.usuario.cedula
    archivo_entrada = os.path.join(settings.MEDIA_ROOT, "plantillas", "plantilla.xlsx")

    if not os.path.exists(archivo_entrada):
        return HttpResponse("ERROR: No existe la plantilla en MEDIA/plantillas/plantilla.xlsx")

    # ---------------------------
    # Función que replica la plantilla por cada cuota
    # ---------------------------
    def replicar_plantilla(archivo_entrada, cuotas, cedula_cliente, credito, prenda):
        wb = openpyxl.load_workbook(archivo_entrada)
        ws = wb.active

        inicio_fila = 1
        fin_fila = 15
        inicio_columna = 1
        fin_columna = 18

        # --- DATOS DEL CLIENTE ---
        user = credito.usuario
        nombre_completo = f"{user.nombre} {user.p_apellido} {user.s_apellido}".strip()

        # --- 🔥 DATOS DE LA PRENDA ---
        nombre_prenda = f"{prenda.nombres} {prenda.apellidos}".strip()

        # --- PRIMERA CUOTA ---
        primera_cuota = cuotas[0]
        fecha = primera_cuota.fecha_vencimiento

        meses_es = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]

        # ------------------------------
        # UBICAR DATOS BASE
        # ------------------------------
        ws.cell(row=1, column=4).value = credito.id
        ws.cell(row=3, column=11).value = nombre_completo
        ws.cell(row=12, column=10).value = user.direccion
        ws.cell(row=12, column=14).value = user.telefono
        ws.cell(row=1, column=6).value = credito.vehiculo.placa

        # 🔥 H7 → NOMBRES Y APELLIDOS DE LA PRENDA
        ws.cell(row=7, column=8).value = nombre_prenda

        # Fecha primera cuota
        ws.cell(row=4, column=10).value = fecha.day
        ws.cell(row=4, column=12).value = meses_es[fecha.month - 1]
        ws.cell(row=4, column=16).value = fecha.year

        ws.cell(row=8, column=11).value = numero_a_letras(primera_cuota.cuota_total)

        ws.cell(row=2, column=15).value = primera_cuota.numero
        ws.cell(row=2, column=17).value = float(primera_cuota.cuota_total)
        ws.cell(row=8, column=17).value = float(primera_cuota.cuota_total)

        c = ws.cell(row=2, column=11)
        c.value = fecha
        c.number_format = "DD/MM/YYYY"

        ws.cell(row=4, column=2).value = f"Céd. o NIT: {cedula_cliente}"

        # ----------------------------------------------
        # GUARDAR PLANTILLA BASE
        # ----------------------------------------------
        celdas_combinadas_originales = [
            r for r in ws.merged_cells.ranges
            if r.min_row >= inicio_fila and r.max_row <= fin_fila
        ]

        plantilla = []
        for fila in ws.iter_rows(
            min_row=inicio_fila,
            max_row=fin_fila,
            min_col=inicio_columna,
            max_col=fin_columna
        ):
            plantilla.append([
                (cell.value, cell.font, cell.fill, cell.border,
                 cell.alignment, cell.number_format)
                for cell in fila
            ])

        # -----------------------------------
        # RÉPLICAS
        # -----------------------------------
        fila_actual = fin_fila + 3

        for cuota in cuotas[1:]:
            for i, fila in enumerate(plantilla):
                for j, (valor, font, fill, border, alignment, number_format) in enumerate(fila):
                    nueva = ws.cell(row=fila_actual + i, column=inicio_columna + j)
                    nueva.value = valor
                    nueva.font = copy(font)
                    nueva.fill = copy(fill)
                    nueva.border = copy(border)
                    nueva.alignment = copy(alignment)
                    nueva.number_format = number_format

            ws.cell(row=fila_actual + 1, column=15).value = cuota.numero
            ws.cell(row=fila_actual + 1, column=17).value = float(cuota.cuota_total)
            ws.cell(row=fila_actual + 7, column=17).value = float(cuota.cuota_total)

            celda_fecha = ws.cell(row=fila_actual + 1, column=11)
            celda_fecha.value = cuota.fecha_vencimiento
            celda_fecha.number_format = "DD/MM/YYYY"

            fecha = cuota.fecha_vencimiento
            ws.cell(row=fila_actual + 3, column=10).value = fecha.day
            ws.cell(row=fila_actual + 3, column=12).value = meses_es[fecha.month - 1]
            ws.cell(row=fila_actual + 3, column=16).value = fecha.year

            ws.cell(row=fila_actual + 7, column=11).value = numero_a_letras(cuota.cuota_total)
            ws.cell(row=fila_actual + 3, column=2).value = f"Céd. o NIT: {cedula_cliente}"

            # 🔄 Combinar celdas
            for rango in celdas_combinadas_originales:
                new_min_row = rango.min_row + fila_actual - inicio_fila
                new_max_row = rango.max_row + fila_actual - inicio_fila
                ws.merge_cells(
                    f"{get_column_letter(rango.min_col)}{new_min_row}:"
                    f"{get_column_letter(rango.max_col)}{new_max_row}"
                )

            fila_actual += len(plantilla) + 2

        temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.seek(0)
        return temp_file

    archivo_temporal = replicar_plantilla(
        archivo_entrada,
        cuotas,
        cedula_cliente,
        credito,
        prenda
    )

    filename = f"Cronograma_{credito.usuario.cedula}_{credito.id}.xlsx"
    return FileResponse(
        open(archivo_temporal.name, "rb"),
        as_attachment=True,
        filename=filename
    )


# -----------------------------------------------------------------------------------
# LISTADO DE CRÉDITOS CON FILTROS Y CONTADOR
# -----------------------------------------------------------------------------------
@login_required
@requiere_sede_financiera_admin
def lista_creditos(request):

    # ---------------------------
    # FILTROS DEL USUARIO
    # ---------------------------
    estado = request.GET.get("estado", "")
    cedula = request.GET.get("cedula", "")
    nombre = request.GET.get("nombre", "")
    placa = request.GET.get("placa", "")
    credito_id = request.GET.get("id", "")
    fecha_desde = request.GET.get("fecha_desde", "")
    fecha_hasta = request.GET.get("fecha_hasta", "")
    valor_min = request.GET.get("valor_min", "")
    valor_max = request.GET.get("valor_max", "")

    # ---------------------------
    # BASE QUERY
    # ---------------------------
    creditos = Credito.objects.all().select_related("usuario", "vehiculo")

    # ---------------------------
    # FILTROS APLICADOS
    # ---------------------------
    
    if estado:
        creditos = creditos.filter(estado=estado)

    if cedula:
        creditos = creditos.filter(usuario__cedula__icontains=cedula)

    if nombre:
        creditos = creditos.filter(
            Q(usuario__nombre__icontains=nombre) |
            Q(usuario__p_apellido__icontains=nombre) |
            Q(usuario__s_apellido__icontains=nombre)
        )

    if placa:
        creditos = creditos.filter(vehiculo__placa__icontains=placa)

    if credito_id:
        creditos = creditos.filter(id=credito_id)

    if fecha_desde:
        creditos = creditos.filter(fecha_creacion__date__gte=fecha_desde)

    if fecha_hasta:
        creditos = creditos.filter(fecha_creacion__date__lte=fecha_hasta)

    if valor_min:
        creditos = creditos.filter(valor_inicial__gte=valor_min)

    if valor_max:
        creditos = creditos.filter(valor_inicial__lte=valor_max)

    # ---------------------------
    # CONTADORES POR ESTADO
    # ---------------------------
    contadores = {
        "Activo": Credito.objects.filter(estado="Activo").count(),
        "En mora": Credito.objects.filter(estado="En mora").count(),
        "Cobro jurídico": Credito.objects.filter(estado="Cobro jurídico").count(),
        "Finalizado": Credito.objects.filter(estado="Finalizado").count(),
        "Total": Credito.objects.count(),
    }

    # ---------------------------
    # PAGINACIÓN
    # ---------------------------
    paginator = Paginator(creditos, 10)  # 10 por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Mantener filtros en paginación
    querystring = "&".join([
        f"{key}={value}"
        for key, value in request.GET.items()
        if key != "page" and value
    ])

    # ---------------------------
    # RENDER
    # ---------------------------
    return render(request, "creditos/lista_creditos.html", {
        "page_obj": page_obj,
        "contadores": contadores,
        "estado_seleccionado": estado,
        "querystring": querystring,
    })

# -----------------------------------------------------------------------------------
# BUSCAR CLIENTE POR CÉDULA + DOCUMENTOS + CRÉDITOS RELACIONADOS
# -----------------------------------------------------------------------------------
@login_required
@requiere_sede_financiera_admin
def cliente_detalle(request):
    cedula = request.GET.get('cedula', '')
    user = None
    creditos = None
    form = None

    if cedula:
        try:
            user = usuario.objects.get(cedula=cedula)
            creditos = user.creditos.all()

            if request.method == 'POST':
                form = DocumentosUsuarioForm(request.POST, request.FILES, instance=user)
                if form.is_valid():
                    form.save()
                    messages.success(request, "Documentos actualizados correctamente.")
            else:
                form = DocumentosUsuarioForm(instance=user)

        except usuario.DoesNotExist:
            messages.error(request, "Usuario no encontrado.")

    return render(request, 'creditos/cliente_detalle.html', {
        'user': user,
        'creditos': creditos,
        'form': form,
        'cedula_buscada': cedula
    })

# -----------------------------------------------------------------------------------
# BUSCAR VEHÍCULO POR PLACA (PRÉNDA + CRÉDITO)
# -----------------------------------------------------------------------------------
@login_required
@requiere_sede_financiera_admin
def buscar_vehiculo(request):
    placa = request.GET.get('placa', '').upper()
    vehiculo = None
    creditos = None
    user = None

    if placa:
        try:
            vehiculo = Vehiculo_contratos.objects.get(placa=placa)
            credito_asociado = vehiculo.creditos.first()
            user = credito_asociado.usuario if credito_asociado else None
            creditos = vehiculo.creditos.all()

        except Vehiculo_contratos.DoesNotExist:
            messages.error(request, "Vehículo no encontrado.")

    return render(request, 'creditos/buscar_vehiculo.html', {
        'vehiculo': vehiculo,
        'user': user,
        'creditos': creditos,
        'placa_buscada': placa,
    })

# ---------------------------------------------------------------
# DASHBOARD PRINCIPAL
# ---------------------------------------------------------------
@login_required
@requiere_sede_financiera_admin
def dashboard(request):
    hoy = timezone.now().date()
    fecha_alerta = hoy + timedelta(days=7)

    # ---------------- METRICAS GENERALES ----------------
    total_creditos = Credito.objects.count()

    total_prestado = Credito.objects.aggregate(total=Sum("valor_inicial"))["total"] or 0
    total_pagado = PagoCredito.objects.aggregate(total=Sum("valor_pagado"))["total"] or 0

    activos = Credito.objects.filter(estado="Activo").count()
    mora = Credito.objects.filter(estado="En mora").count()
    juridico = Credito.objects.filter(estado="Cobro jurídico").count()
    finalizados = Credito.objects.filter(estado="Finalizado").count()

    recuperacion = 0
    if total_prestado > 0:
        recuperacion = (total_pagado / total_prestado) * 100

    # ---------------- VALOR TOTAL EN MORA (sumando cuotas en Mora) ----------------
    cuotas_en_mora_qs = CuotaCredito.objects.filter(estado="Mora")
    valor_mora = cuotas_en_mora_qs.aggregate(total=Sum("cuota_total"))["total"] or 0

    # Creditos al dia (simplemente los activos)
    al_dia = activos

    cartera_activa = total_prestado - total_pagado
    cartera_vencida = valor_mora

    # ---------------- ALERTAS ----------------
    # Cuotas próximas: cuotas NO pagadas con fecha_vencimiento entre hoy y fecha_alerta
    cuotas_proximas_qs = CuotaCredito.objects.filter(
        pagada=False,
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=fecha_alerta
    ).select_related("credito").order_by("fecha_vencimiento")[:10]

    # Construimos una lista amigable para template (contiene cuota, credito, fecha_venc)
    cuotas_proximas = [{
        "credito": cuota.credito,
        "cuota_numero": cuota.numero,
        "fecha_vencimiento": cuota.fecha_vencimiento,
        "cuota_total": cuota.cuota_total
    } for cuota in cuotas_proximas_qs]

    # Créditos en mora > 30 días (buscamos créditos que tengan alguna cuota en mora con fecha_vencimiento <= hoy-30)
    fecha_30 = hoy - timedelta(days=30)
    creditos_mora_prolongada_qs = Credito.objects.filter(
        cuotas_credito__estado="Mora",
        cuotas_credito__fecha_vencimiento__lte=fecha_30
    ).distinct().select_related("usuario")[:10]

    mora_mas_30 = list(creditos_mora_prolongada_qs)

    # ---------------- GRAFICAS ----------------
    # Créditos por estado
    estados_labels = ["Activos", "En mora", "Cobro jurídico", "Finalizados"]
    estados_data = [activos, mora, juridico, finalizados]

    # Pagos mensuales (por fecha_pago)
    pagos_mensuales_qs = (
        PagoCredito.objects
        .annotate(mes=TruncMonth("fecha_pago"))
        .values("mes")
        .annotate(total=Sum("valor_pagado"))
        .order_by("mes")
    )
    pagos_labels = [p["mes"].strftime("%b %Y") for p in pagos_mensuales_qs]
    pagos_data = [float(p["total"]) for p in pagos_mensuales_qs]

    # Créditos otorgados por mes (por fecha_inicio)
    creditos_mensuales_qs = (
        Credito.objects
        .annotate(mes=TruncMonth("fecha_inicio"))
        .values("mes")
        .annotate(total=Sum("valor_inicial"))
        .order_by("mes")
    )
    creditos_labels = [c["mes"].strftime("%b %Y") for c in creditos_mensuales_qs]
    creditos_data = [float(c["total"]) for c in creditos_mensuales_qs]

    # Mora mensual: sumar cuota_total de cuotas en Mora por mes de fecha_vencimiento
    mora_mensual_qs = (
        CuotaCredito.objects.filter(estado="Mora")
        .annotate(mes=TruncMonth("fecha_vencimiento"))
        .values("mes")
        .annotate(total=Sum("cuota_total"))
        .order_by("mes")
    )
    mora_labels = [m["mes"].strftime("%b %Y") for m in mora_mensual_qs]
    mora_data = [float(m["total"]) for m in mora_mensual_qs]

    # ---------------- ULTIMOS REGISTROS ----------------
    ultimos_pagos = PagoCredito.objects.select_related("credito").order_by("-id")[:10]
    ultimos_creditos = Credito.objects.select_related("usuario").order_by("-id")[:10]
    

    # ---------------- CONTEXTO ----------------
    context = {
        # métricas
        "total_creditos": total_creditos,
        "total_prestado": float(total_prestado),
        "total_pagado": float(total_pagado),
        "activos": activos,
        "mora": mora,
        "juridico": juridico,
        "finalizados": finalizados,
        "recuperacion": round(recuperacion, 2),
        "valor_mora": float(valor_mora),
        "al_dia": al_dia,
        "cartera_activa": float(cartera_activa),
        "cartera_vencida": float(cartera_vencida),

        # alertas
        "cuotas_proximas": cuotas_proximas,
        "mora_mas_30": mora_mas_30,

        # graficas
        "estados_labels": estados_labels,
        "estados_data": estados_data,
        "pagos_labels": pagos_labels,
        "pagos_data": pagos_data,
        "creditos_labels": creditos_labels,
        "creditos_data": creditos_data,
        "mora_labels": mora_labels,
        "mora_data": mora_data,

        # listas
        "ultimos_pagos": ultimos_pagos,
        "ultimos_creditos": ultimos_creditos,
    }

    return render(request, "dashboard/dashboard.html", context)


@login_required
@requiere_sede_financiera_admin
def renovar_servicio(request, credito_id):
    credito = get_object_or_404(Credito, id=credito_id)

    if request.method == 'POST':
        form = RenovarServicioForm(request.POST)
        if form.is_valid():
            tipo = form.cleaned_data['tipo']
            valor_anual = form.cleaned_data['valor_anual']
            fecha_inicio = form.cleaned_data['fecha_inicio']
            fecha_fin = fecha_inicio + relativedelta(months=12) - relativedelta(days=1)

            # si existe un registro pendiente para ese tipo y fecha_inicio, activarlo
            servicio, created = ServicioCredito.objects.get_or_create(
                credito=credito,
                tipo=tipo,
                fecha_inicio=fecha_inicio,
                defaults={'fecha_fin': fecha_fin, 'valor_anual': valor_anual, 'estado': ServicioCredito.ESTADO_ACTIVO}
            )
            if not created:
                servicio.valor_anual = valor_anual
                servicio.estado = ServicioCredito.ESTADO_ACTIVO
                servicio.fecha_fin = fecha_fin
                servicio.save()

            messages.success(request, f"{dict(ServicioCredito.TIPOS).get(tipo)} renovado y activado desde {fecha_inicio}.")
            # opcional: regenerar cronograma (recalcular cuotas futuras)
            try:
                credito.generar_cronograma(recalcular=True)
            except ValueError as e:
                messages.warning(request, f"Renovado pero persisten pendientes: {e}")

            return redirect('detalle_credito', credito_id=credito.id)
    else:
        # Por defecto, ofrecer renovar el siguiente periodo pendiente más cercano
        proximo_inicio = None
        # buscar primer servicio pendiente para cualquier tipo
        pendiente = credito.servicios.filter(estado=ServicioCredito.ESTADO_PENDIENTE).order_by('fecha_inicio').first()
        if pendiente:
            proximo_inicio = pendiente.fecha_inicio

        form = RenovarServicioForm(initial={'fecha_inicio': proximo_inicio})
    return render(request, 'creditos/renovar_servicio.html', {'form': form, 'credito': credito})

def numero_a_letras(valor):
    try:
        texto = num2words(int(valor), lang='es')
        return texto.upper() + " PESOS"
    except:
        return ""

@login_required
@requiere_sede_financiera_admin
def generar_recibo(request):
    ids = request.GET.get("ids", "")
    ids_list = [int(x) for x in ids.split(",") if x.isdigit()]

    pagos = PagoCredito.objects.filter(id__in=ids_list)
    if not pagos.exists():
        return HttpResponse("No se encontraron pagos.", status=404)

    credito = pagos.first().credito
    fecha_hora = timezone.now()
    total_operacion = sum(p.valor_pagado for p in pagos)

    pagos_items = []

    for p in pagos:
        item = {
            "tipo_pago": p.tipo_pago,
            "cuota_numero": p.cuota_numero,
            "valor_pagado": p.valor_pagado,
            "observacion": p.observacion,
            "interes_mora": p.interes_moratorio_pagado,
            "cobro_juridico": p.cobro_juridico_pagado,
            "dias_mora": 0,
        }

        if p.tipo_pago == "cuota" and p.cuota_numero:
            cuota = credito.cuotas_credito.filter(numero=p.cuota_numero).first()
            if cuota:
                mora_info = credito.calcular_moratorios_por_cuota(cuota)
                item["dias_mora"] = mora_info.get("dias", 0)

        pagos_items.append(item)

    context = {
        "pago_operacion": {
            "consecutivo": pagos.first().id,
            "items_count": len(pagos_items),
            "fecha_hora": fecha_hora,
        },
        "pagos_items": pagos_items,
        "total_operacion": total_operacion,
        "credito": credito,
    }

    # -------------------------
    # 🔥 GENERAR PDF
    # -------------------------
    html_string = render_to_string("creditos/recibo_pago.html", context)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    # -------------------------
    # 🔥 GUARDAR RECIBO
    # -------------------------
    recibo = ReciboPago.objects.create(
        credito=credito,
        total=total_operacion
    )
    recibo.pagos.set(pagos)

    nombre_archivo = f"recibo_credito_{credito.id}_{recibo.id}.pdf"
    recibo.archivo.save(nombre_archivo, ContentFile(pdf_file))
    recibo.save()

    # -------------------------
    # 🔥 DEVOLVER PDF
    # -------------------------
    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{nombre_archivo}"'
    return response


@login_required
@requiere_sede_financiera_admin
def solicitudes_credito(request):
    solicitudes = SolicitudCredito.objects.all().order_by('-id')
    return render(request, "creditos/solicitudes.html", {"solicitudes": solicitudes})

@login_required
@requiere_sede_financiera_admin
def cambiar_estado_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudCredito, id=solicitud_id)
    prendas = NombrePrendas.objects.all()

    if request.method == "POST":
        nuevo_estado = request.POST.get("estado")
        prenda_id = request.POST.get("prenda")
        valor_aprobado = request.POST.get("valor_aprobado")
        fecha_aprobacion = request.POST.get("fecha_aprobacion")

        solicitud.estado = nuevo_estado

        # ✅ SI SE APRUEBA
        if nuevo_estado == 'Aprobado':

            # FECHA DE APROBACIÓN MANUAL
            if fecha_aprobacion:
                solicitud.fecha_aprobacion = fecha_aprobacion
            else:
                solicitud.fecha_aprobacion = timezone.now().date()

            # VALOR APROBADO
            try:
                solicitud.valor_aprobado = float(valor_aprobado)
            except (TypeError, ValueError):
                solicitud.valor_aprobado = solicitud.valor

            # PRENDA
            if prenda_id:
                solicitud.prenda = get_object_or_404(NombrePrendas, id=prenda_id)
            else:
                solicitud.prenda = None

        # ❌ SI NO ESTÁ APROBADO
        else:
            solicitud.fecha_aprobacion = None
            solicitud.valor_aprobado = None
            solicitud.prenda = None

        solicitud.save()
        messages.success(request, "Estado de la solicitud actualizado correctamente.")
        return redirect("solicitudes_credito")

    return render(request, "creditos/cambiar_estado_solicitud.html", {
        "solicitud": solicitud,
        "prendas": prendas
    })


def generar_pdf_aprobacion(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudCredito, id=solicitud_id)

    if solicitud.estado != "Aprobado":
        messages.error(request, "Solo las solicitudes aprobadas pueden generar PDF.")
        return redirect("solicitudes_credito")

    contexto = {
        "solicitud": solicitud,
        "usuario": solicitud.usuario,
        "vehiculo": solicitud.vehiculo,
        "prenda": solicitud.prenda,

        # 👉 NUEVOS DATOS
        "valor_aprobado": solicitud.valor_aprobado,
        "fecha_aprobacion": solicitud.fecha_aprobacion,
    }

    html = render_to_string("creditos/aprobacion.html", contexto)
    return generar_pdf(html, f"aprobacion_{solicitud.id}.pdf")


def generar_pdf_prenda(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudCredito, id=solicitud_id)

    if solicitud.estado != 'Aprobado':
        messages.error(
            request,
            "No puedes generar la prenda hasta que la solicitud sea aprobada."
        )
        return redirect('solicitudes_credito')

    if not solicitud.prenda:
        messages.error(
            request,
            "Esta solicitud no tiene una persona de prenda asociada."
        )
        return redirect('solicitudes_credito')

    context = {
        'solicitud': solicitud,
        'user': solicitud.usuario,
        'vehiculo': solicitud.vehiculo,
        'prenda': solicitud.prenda,

        # 👉 NUEVOS CAMPOS
        'valor_aprobado': solicitud.valor_aprobado,
        'fecha_aprobacion': solicitud.fecha_aprobacion,
    }

    html_string = render_to_string(
        'creditos/contrato_prenda.html',
        context
    )

    html = HTML(
        string=html_string,
        base_url=request.build_absolute_uri()
    )
    pdf = html.write_pdf()

    filename = f"Prenda_{solicitud.prenda.documento}_{solicitud.vehiculo.placa}.pdf"
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    return response


def generar_pdf(html_string, filename):
    pdf = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

@login_required
def crear_vehiculo_credito(request):
    if request.method == "POST":
        form = VehiculoCreditoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Vehículo creado correctamente.")
            return redirect("crear_credito")
    else:
        form = VehiculoCreditoForm()

    return render(request, "creditos/crear_vehiculo_credito.html", {"form": form})

@login_required
def crear_usuario_credito(request):
    if request.method == "POST":
        form = UsuarioCreditoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario creado correctamente.")
            return redirect("crear_credito")  # vuelve al formulario de crédito
    else:
        form = UsuarioCreditoForm()

    return render(request, "creditos/crear_usuario_credito.html", {"form": form})

@login_required
def creditos_por_estado(request, estado):
    creditos = Credito.objects.filter(estado=estado).select_related("usuario")

    context = {
        "creditos": creditos,
        "estado": estado,
    }
    return render(request, "creditos/creditos_por_estado.html", context)

@login_required
def crear_prenda(request):
    if request.method == 'POST':
        form = NombrePrendasForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Prenda creada correctamente.')
            return redirect('crear_prenda')
    else:
        form = NombrePrendasForm()

    return render(request, 'creditos/crear_prenda.html', {
        'form': form
    })